#-------------------------------------------------------------------------------
# Name:        Cleaning up Duplicates
# Author:      Kame Hizzouse
# License:     MIT License
#-------------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import io
import re
import unicodedata


import tkinter
from tkinter import filedialog
import html5lib
import time
import random


#remember to have requests, BeautifulSoup4, requests[security], and in your packages for this to work!
import requests
from bs4 import BeautifulSoup
#Define starting constants:

RowCounter = 2
#Beginning and Ending Cells in your worksheet to Process
StartCell = 2
FinalCell = 1943

def re_list(a_list, a_str):
    sub = ''
    sub = ''.join(a_list)
    sub = sub.split(a_str)    
    return sub

#Open the file in question:

open_file = filedialog.askopenfilename ()

#Dialog Box opening complete

print("Now opening: " + str(open_file))

#Performing initial IO on workbooks
wb = load_workbook(filename = open_file)
blank = Workbook()
ws = wb.active

DrugsInput = wb["Drugs"]
Results = wb["TimeStats"]
TrialCounter = StartCell
FinishedLinesTest = []
while(TrialCounter<4):

    DN = DrugsInput['A' + str(TrialCounter)].value
    
    RawLines = io.StringIO(str(DN))
    
    ActualLines = RawLines.readlines()
    ProperSyntax = [x.strip() for x in ActualLines]
    
    ProperSyntax = ''.join(ProperSyntax)
    ProperSyntax = str(ProperSyntax)
    print(ProperSyntax)
    Results["A" + str(TrialCounter)] = ProperSyntax
    
    GS = DrugsInput['C' + str(TrialCounter)].value
    
    GSTag = io.StringIO(str(GS))
    GlobalStatus = GSTag.readline()
    #Read the Excel file and put into array
    
    
    RowPilot = 0
    
    Results["B" + str(RowCounter)] = GlobalStatus


    KCPC = DrugsInput['BF' + str(TrialCounter)].value
    PreclinicalTrial = io.StringIO(str(KCPC))

    PreclinicalTrialData = PreclinicalTrial.readlines()
    print(PreclinicalTrialData)
    #if(PreclinicalTrialData = [] or )
    KCPCInfo = [x.strip() for x in PreclinicalTrialData]
    print(KCPCInfo)
    if(KCPCInfo == ['None'] or KCPCInfo == ['-']):
        TrialCounter += 1
        RowCounter += 1
        continue

    KCPCInfo = re_list(KCPCInfo, ",")
    print(KCPCInfo)
    
    
    test = 'abc3'
    print(test.isdigit()) 
    print(test.isalnum())
    print(test.isnumeric())

    TrialCounter += 1    

wb.save(open_file)
print("Done")







