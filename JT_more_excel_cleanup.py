#-------------------------------------------------------------------------------
# Name:        More Excel File Cleanup
# Author:      Jonathan Thomas
#
# Licence:     MIT License
#-------------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import StringIO
import re
import unicodedata

import Tkinter
import tkFileDialog
import html5lib
import time
import random


#remember to have requests, BeautifulSoup4, requests[security], and in your packages for this to work!
import requests
from bs4 import BeautifulSoup
#Define starting constants:

#Beginning and Ending Cells in your worksheet to Process
StartCell = 2
EndCell = 2927
#Code Shamelessly copied from StackExchange user HennyH
def search(text,n):
    '''Searches for text, and retrieves n words either side of the text, which are returned seperately'''
    word = r"\W*([\w]+)"
    groups = re.search(r'{}\W*{}{}'.format(word*n,'place',word*n), text).groups()
    return groups[:n],groups[n:]

def find_all(a_str, sub):
    start = 0
    while True:
        start = a_str.find(sub, start)
        if start == -1: return
        yield start
        start += len(sub) # use start += 1 to find overlapping matches


def find_between_r( s, first, last ):
    try:
        start = s.rindex( first ) + len( first )
        end = s.rindex( last, start )
        return s[start:end]
    except ValueError:
        return ""

#Open the file in question:

open_file = tkFileDialog.askopenfilename ()

#Dialog Box opening complete

print("Now2 opening: " + str(open_file))

#Performing initial IO on workbooks
wb = load_workbook(filename = open_file)
blank = Workbook()
ws = wb.active

ResultsInput = wb["TheDrugz"]
FinishedLinesCheck = []
CellIter = StartCell
while(CellIter<=EndCell):
    print(CellIter)
    OBJ = ResultsInput['E' + str(CellIter)].value

    RawLines = StringIO.StringIO(str(OBJ))


    #Read the Excel file and put into array
    ActualLines = RawLines.readlines()

    FinishedLines = [x.strip() for x in ActualLines]
    
    StartDate = FinishedLines[0]
    print(FinishedLines[0])
    
    OBJ = ResultsInput['F' + str(CellIter)].value

    RawLines = StringIO.StringIO(str(OBJ))

    #Read the Excel file and put into array
    ActualLines = RawLines.readlines()

    FinishedLines = [x.strip() for x in ActualLines]
    CompletionDate = FinishedLines[0]

    CompOne = StartDate
    CompTwo = CompletionDate
    Months = [CompOne, CompTwo]
    Years = [CompOne, CompTwo]
    months = [CompOne, CompTwo]
    Value = 0
    
    k = 0
    while(k < 2):
        if(Months[k][0:3]=='Jan'):
            Months[k] = 1
        elif(Months[k][0:3]=='Feb'):
            Months[k] = 2
        elif(Months[k][0:3]=='Mar'):
            Months[k] = 3
        elif(Months[k][0:3]=='Apr'):
            Months[k] = 4
        elif(Months[k][0:3]=='May'):
            Months[k] = 5
        elif(Months[k][0:3]=='Jun'):
            Months[k] = 6
        elif(Months[k][0:3]=='Jul'):
            Months[k] = 7
        elif(Months[k][0:3]=='Aug'):
            Months[k] = 8
        elif(Months[k][0:3]=='Sep'):
            Months[k] = 9
        elif(Months[k][0:3]=='Oct'):
            Months[k] = 10
        elif(Months[k][0:3]=='Nov'):
            Months[k] = 11
        elif(Months[k][0:3]=='Dec'):
            Months[k] = 12
        else:
            Months[k] = 0
        k = k + 1
            
    l = 0
    while(l < 2):
        if(months[l][0:3]=='Jan'):
            months[l] = .01
        elif(months[l][0:3]=='Feb'):
            months[l] = .02
        elif(months[l][0:3]=='Mar'):
            months[l] = .03
        elif(months[l][0:3]=='Apr'):
            months[l] = .04
        elif(months[l][0:3]=='May'):
            months[l] = .05
        elif(months[l][0:3]=='Jun'):
            months[l] = .06
        elif(months[l][0:3]=='Jul'):
            months[l] = .07
        elif(months[l][0:3]=='Aug'):
            months[l] = .08
        elif(months[l][0:3]=='Sep'):
            months[l] = .09
        elif(months[l][0:3]=='Oct'):
            months[l] = .10
        elif(months[l][0:3]=='Nov'):
            months[l] = .11
        elif(months[l][0:3]=='Dec'):
            months[l] = .12
        else:
            months[l] = .0
        l = l + 1
        
    m = 0
    while(m < 2):
        Years[m] = Years[m][-4:]
        print(Years[m].isdigit())
        if(Years[m].isdigit()):
            Years[m] = int(Years[m])
        m += 1
     
    print(Months)
    print(Years)
    print(months)    
    
    print(isinstance(Months[0], int) and isinstance(Years[1], int))
    
    if(isinstance(Months[0], int) and isinstance(Years[1], int)):
        Value = int(Years[1]) + (months[1])
        print(Value)
            
    if(Value > 2018.02):
        Years[1] = 2018
        Months[1] = 2

    LengthTrial = 0
    
    print(Years[0])
    print(Years[1])

    print(isinstance(Years[0], int) and isinstance(Years[1], int))
    
    if(isinstance(Years[0], int) and isinstance(Years[1], int)):
        if(Years[1] - Years[0] > 1):
#           print(Years[0], Years[1], Months[0], Months[1])
            LengthTrial = ((12 - Months[0]) + Months[1]) + 12*((Years[1] - Years[0]) - 1)
        elif(Years[1] - Years[0] == 1):
#           print(Years[0], Years[1], Months[0], Months[1])
            LengthTrial = ((12 - Months[0]) + Months[1])
        elif(Years[1] - Years[0] == 0):
#           print(Years[0], Years[1], Months[0], Months[1])
            LengthTrial = int(Months[1]) - int(Months[0])
        else:
#           print(Years[0], Years[1], Months[0], Months[1])
            LengthTrial = 0        
        if(LengthTrial < 0):
            LengthTrial = abs(LengthTrial)
    else:
        LengthTrial = 'Not Available'
            
    ResultsInput['G' + str(CellIter)] = LengthTrial
                        
    print("Finished Scanning Cell: " + str(CellIter) + " of " + str(EndCell)) + "."
    CellIter = CellIter + 1

wb.save(open_file)
print("Done")