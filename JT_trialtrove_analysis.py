#-------------------------------------------------------------------------------
# Name:        TrialTrove Data Anlysis
# Purpose:     Fills in a separate sheet next to the TrialTrove Excel data,
#              calculates Trial Lengths, and organizes the information.
#
# Author:      Jonathan Thomas
# 
# License:     MIT "Free Beer" License
#-------------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.utils import get_column_letter
import StringIO
import re
import unicodedata

import Tkinter
import tkFileDialog
import html5lib
import time
import random


#Beginning and Ending Cells in your worksheet to Process
StartCell = 2
EndCell = 3323

#Open the file in question:

open_file = tkFileDialog.askopenfilename ()

#Dialog Box opening complete

print("Now opening: " + str(open_file))

#Performing initial IO on workbooks
wb = load_workbook(filename = open_file)
blank = Workbook()
ws = wb.active

sheet_ranges = wb["trialtrove_4855441"]
ResultsInput = wb["Time"]
InputRowCounter = StartCell
TTncts = []
ttncts = []

while(InputRowCounter != EndCell):   
    if(InputRowCounter == 1000):
        wb.save(open_file)
    if(InputRowCounter == 2000):
        wb.save(open_file)
    if(InputRowCounter == 3000):
        wb.save(open_file)
        
    OBJ1 = sheet_ranges['J' + str(InputRowCounter)].value
    if(OBJ1 is None):
        InputRowCounter += 1
        continue
    
    PCSubjects = re.findall(r"NCT\d\d\d\d\d\d\d\d", OBJ1)
    if(PCSubjects == [] or PCSubjects == ['-']):
        InputRowCounter += 1
        continue
    
    i = 0
    while(i<len(PCSubjects)):
        PCSubjects[i] = PCSubjects[i].encode('utf-8')
        i += 1
        
    TTncts.extend(PCSubjects)

    InputRowCounter += 1    

if(InputRowCounter == EndCell):
    ttncts = list(set(TTncts))
    
Sum1 = ''
Sum2 = ''
Summary = ''
StudyType = ''
Phases = ''
Age1 = ''
Age2 = ''
Age3 = ''
Age4 = ''
Ages = ''
Drugs1 = ''
Drugs2 = ''
Drugs = ''
Identifier = ''
Criteria = ''
TimeofStudy = ''
DetailedDescription = ''
StartDate = ''
CompletionDate = ''

InputRowCounter = 2
ResultsRowCounter = 2 
while(InputRowCounter != EndCell):    
    print(InputRowCounter)
    print(ResultsRowCounter)   
    StartDate = ''
    CompletionDate = ''
    OBJ = sheet_ranges['J' + str(InputRowCounter)].value
    if(OBJ is None):
        InputRowCounter += 1
        continue
    
    PCSubjects = re.findall(r"NCT\d\d\d\d\d\d\d\d", OBJ)
    if(PCSubjects == [] or PCSubjects == ['-']):
        InputRowCounter += 1
        continue
    print(PCSubjects[0])
    Identifier = str(PCSubjects[0]) 
    
    OBJ = sheet_ranges['C' + str(InputRowCounter)].value
    Phase = OBJ
    
    OBJ = sheet_ranges['H' + str(InputRowCounter)].value
    Drugs1 = OBJ    
    
    OBJ = sheet_ranges['I' + str(InputRowCounter)].value
    Drug2 = OBJ    

    Drugs = Drugs1 + ' , ' + Drugs2
    
    
    OBJ = sheet_ranges['AI' + str(InputRowCounter)].value
    RawLines = StringIO.StringIO(OBJ)
    print(OBJ)
    #Read the Excel file and put into array
    ActualLines = RawLines.readlines()
    FinishedLines = [x.strip() for x in ActualLines]
    print(FinishedLines)
    FinishedLines[0] = FinishedLines[0].encode('utf-8')
    print(FinishedLines[0])
    StudyType = FinishedLines[0]
    
    OBJ = sheet_ranges['A' + str(InputRowCounter)].value
    Sum1 = OBJ
    OBJ = sheet_ranges['L' + str(InputRowCounter)].value
    Sum2 = OBJ
    
    Summary = Sum1 + ' , ' + Sum2
    
    OBJ = sheet_ranges['Q' + str(InputRowCounter)].value
    Sum1 = OBJ    
    OBJ = sheet_ranges['R' + str(InputRowCounter)].value
    Sum2 = OBJ  
    
    Criteria = Sum1 + ' , ' + Sum2
    
    OBJ = sheet_ranges['U' + str(InputRowCounter)].value
    Age1 = OBJ    
    OBJ = sheet_ranges['V' + str(InputRowCounter)].value
    Age2 = OBJ
    OBJ = sheet_ranges['W' + str(InputRowCounter)].value
    Age3 = OBJ    
    OBJ = sheet_ranges['X' + str(InputRowCounter)].value
    Age4 = OBJ    
    
    Ages = Age1 + ' , ' + Age2 + ' , ' + Age3 + ' , ' + Age4
    
    OBJ = sheet_ranges['O' + str(InputRowCounter)].value
    RawLines = StringIO.StringIO(OBJ)
    ActualLines = RawLines.readlines()
    FinishedLines = [x.strip() for x in ActualLines]
    if(FinishedLines != [] and FinishedLines != ['None']):   
        StartDate = FinishedLines[0]  
        Months1 = int(FinishedLines[0][5:7])
        Years1 = int(FinishedLines[0][0:4])
    else:
        Months1 = 0
        Years1 = 0
        StartDate = 'No Start Date Reported'
    
    OBJ = sheet_ranges['P' + str(InputRowCounter)].value
    RawLines = StringIO.StringIO(OBJ)
    ActualLines = RawLines.readlines()
    FinishedLines = [x.strip() for x in ActualLines] 
    if(FinishedLines != [] and FinishedLines != ['None']):    
        CompletionDate = FinishedLines[0]
        Months2 = int(FinishedLines[0][5:7])
        Years2 = int(FinishedLines[0][0:4])
    else:
        Months2 = 0
        Years2 = 0
        CompletionDate = 'No Completion Date Reported'
    
    if(isinstance(Years1, (int, long)) and isinstance(Years2, (int, long))):
        if(Years1 == 0 or Years2 == 0):
            TimeOfStudy = 'N/A'
        elif((Years2 - Years1) > 1):
            TimeOfStudy = str(((12 - Months1) + Months2) + 12*(int(Years2) - int(Years1) - 1))
        elif(int(Years2) - int(Years1) == 1):
            TimeOfStudy = ((12 - Months1) + Months2)
        elif(int(Years2) - int(Years1) == 0):
            TimeOfStudy = int(Months2) - int(Months1)
        else:
            TimeOfStudy = 0
        if(TimeOfStudy < 0):
            TimeOfStudy = abs(TimeOfStudy)
    else:
        TimeOfStudy = 'Not Available'
    
    ResultsInput['A' + str(ResultsRowCounter)] = Identifier
    ResultsInput['B' + str(ResultsRowCounter)] = Drugs
    ResultsInput['C' + str(ResultsRowCounter)] = Phase
    ResultsInput['D' + str(ResultsRowCounter)] = TimeOfStudy
    ResultsInput['E' + str(ResultsRowCounter)] = StartDate
    ResultsInput['F' + str(ResultsRowCounter)] = CompletionDate
    ResultsInput['G' + str(ResultsRowCounter)] = StudyType
    ResultsInput['H' + str(ResultsRowCounter)] = Summary
    ResultsInput['I' + str(ResultsRowCounter)] = Ages            
    ResultsInput['J' + str(ResultsRowCounter)] = Criteria

    print("Finished Scanning Cell: " + str(InputRowCounter) + " of " + str(EndCell)) + ". Rows of data Filled: " + str(ResultsRowCounter) + "."
    ResultsRowCounter = ResultsRowCounter + 1
    InputRowCounter = InputRowCounter + 1

wb.save(open_file)
print("Done")








